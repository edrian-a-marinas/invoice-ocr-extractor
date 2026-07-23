import os
import openpyxl
from app.core.schemas import ReceiptRecord

HEADERS = [
    "Page",
    "Receipt No.",
    "Doctor Name",
    "PRC License",
    "Hospital",
    "Date",
    "Patient Name",
    "Total Amount (PHP)",
    "Signature",
]
CENTERED_COLUMNS = {1, 2, 6, 8, 9}  # Page, Receipt No., Date, Total Amount, Signature


def write_records_to_excel(records: list[ReceiptRecord], output_path: str) -> None:
    center = openpyxl.styles.Alignment(horizontal="center")

    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        next_row = ws.max_row + 1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            if col in CENTERED_COLUMNS:
                cell.alignment = center
        next_row = 2

    for offset, record in enumerate(records):
        row_idx = next_row + offset
        values = [
            record.page,
            record.receipt_no,
            record.doctor_name,
            record.prc_license,
            record.hospital,
            record.date,
            record.patient_name,
            record.total_amount,
            record.signature,
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col == 8:
                cell.number_format = '"PHP "#,##0.00'
            if col in CENTERED_COLUMNS:
                cell.alignment = center

    for col_idx, header in enumerate(HEADERS, start=1):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        max_length = len(header)
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[column_letter].width = max_length + 4

    wb.save(output_path)
