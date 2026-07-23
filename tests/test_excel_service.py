import openpyxl

from app.core.schemas import ReceiptRecord
from app.services.excel_service import write_records_to_excel, HEADERS


def _make_record(**overrides):
    defaults = dict(
        page=1,
        receipt_no="OR-65116",
        doctor_name="Dr. Graciano Lopez Jaena",
        prc_license="PRC Lic. No. 00010",
        hospital="QUIRINO MEMORIAL MEDICAL CENTER",
        date="August 25, 2024",
        patient_name="Natividad Soriano Reyes",
        total_amount=3900.0,
        signature="Yes",
    )
    defaults.update(overrides)
    return ReceiptRecord(**defaults)


class TestWriteRecordsToExcel:
    def test_creates_file_with_correct_headers(self, tmp_path):
        output_path = tmp_path / "test.xlsx"
        write_records_to_excel([_make_record()], str(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        actual_headers = [ws.cell(row=1, column=col).value for col in range(1, len(HEADERS) + 1)]
        assert actual_headers == HEADERS

    def test_writes_correct_row_values(self, tmp_path):
        output_path = tmp_path / "test.xlsx"
        record = _make_record(receipt_no="OR-99999", patient_name="Juan Dela Cruz")
        write_records_to_excel([record], str(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "OR-99999"
        assert ws.cell(row=2, column=7).value == "Juan Dela Cruz"
        assert ws.cell(row=2, column=8).value == 3900.0

    def test_second_run_appends_not_overwrites(self, tmp_path):
        output_path = tmp_path / "test.xlsx"
        write_records_to_excel([_make_record(receipt_no="OR-11111")], str(output_path))
        write_records_to_excel([_make_record(receipt_no="OR-22222")], str(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        # header row + 2 data rows = 3 rows total
        assert ws.max_row == 3
        assert ws.cell(row=2, column=2).value == "OR-11111"
        assert ws.cell(row=3, column=2).value == "OR-22222"

    def test_empty_records_on_existing_file_does_not_wipe_data(self, tmp_path):
        output_path = tmp_path / "test.xlsx"
        write_records_to_excel([_make_record(receipt_no="OR-11111")], str(output_path))
        write_records_to_excel([], str(output_path))

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        assert ws.max_row == 2  # header + 1 data row, unchanged
        assert ws.cell(row=2, column=2).value == "OR-11111"